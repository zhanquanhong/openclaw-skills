"""
TaskDecomposer - 智能任务分解器

使用 LLM 动态分解复杂任务为自包含的子任务，结构化输出（Pydantic 模型）。
参考：camel/societies/workforce/workforce.py:_decompose_task()
License: Apache 2.0 (CAMEL-AI.org)

v3.4 新增：甘特图生成集成资源配置，支持 Excel 导出
"""

import logging
import re
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any, Tuple

from pydantic import BaseModel, Field  # type: ignore

from ...core.agent import Agent
from ...core.task import Task
from ...core.agent_executor import AgentExecutor
from ...planner import AdaptivePlanner
from .prompts import TASK_DECOMPOSE_PROMPT

logger = logging.getLogger(__name__)


class SubtaskDefinition(BaseModel):
    """子任务定义（结构化输出）"""
    description: str = Field(description="子任务描述，必须是自包含的")
    expected_output: str = Field(description="期望的输出格式或内容")
    dependencies: List[str] = Field(default_factory=list, description="依赖的其他子任务ID列表")
    priority: int = Field(default=1, description="优先级（1-10，数字越大优先级越高）")
    estimated_time: Optional[float] = Field(default=None, description="预估执行时间（秒）")


class TaskDecompositionResult(BaseModel):
    """任务分解结果（结构化输出）"""
    subtasks: List[SubtaskDefinition] = Field(description="子任务列表")
    reasoning: str = Field(description="分解理由")
    can_parallelize: bool = Field(default=True, description="是否可以并行执行")


class ResourceConfig(BaseModel):
    """资源配置（用于甘特图演算）"""
    backend_count: Optional[int] = Field(default=None, description="后端开发人数")
    backend_availability: Optional[float] = Field(default=100.0, description="后端可用时间百分比（0-100）")
    test_count: Optional[int] = Field(default=None, description="测试人数")
    test_availability: Optional[float] = Field(default=100.0, description="测试可用时间百分比（0-100）")
    frontend_count: Optional[int] = Field(default=0, description="前端开发人数")
    product_count: Optional[int] = Field(default=0, description="产品人数")
    target_date: Optional[str] = Field(default=None, description="期望完工日期（YYYY-MM-DD）")
    milestones: Optional[List[str]] = Field(default=None, description="固定里程碑列表")
    skill_level: Optional[str] = Field(default="average", description="团队技能水平：average/senior/mixed")


class TaskDecomposer:
    """任务分解器
    
    使用 LLM 动态分解复杂任务为自包含的子任务，结构化输出。
    """
    
    def __init__(
        self,
        task_agent: Agent,
        llm_provider,
        planner: Optional[AdaptivePlanner] = None,
    ):
        """
        初始化 TaskDecomposer
        
        Args:
            task_agent: Task Agent 实例（负责任务分解）
            llm_provider: LLM 提供者
            planner: AdaptivePlanner 实例（可选，用于任务分解优化）
        """
        self.task_agent = task_agent
        self.llm_provider = llm_provider
        self.planner = planner
        self.resource_config: Optional[ResourceConfig] = None
        
        # 创建 AgentExecutor（启用上下文编译）
        self.executor = AgentExecutor(
            llm_provider=llm_provider,
            enable_context_compilation=True
        )
    
    async def decompose_task(
        self,
        task: Task,
        available_workers: List[Any],  # List[Worker]
        additional_info: Optional[str] = None,
    ) -> List[Task]:
        """
        分解任务为子任务（结构化输出）
        
        Args:
            task: 要分解的任务
            available_workers: 可用的 Worker 列表
            additional_info: 额外信息（可选）
            
        Returns:
            子任务列表（Task 对象）
        """
        logger.info(f"[TaskDecomposer] Decomposing task: {task.id}")
        
        # 构建 Worker 信息字符串
        child_nodes_info = "\n".join([
            f"{worker.id}: {worker.description}: {worker.agent.role}"
            for worker in available_workers
        ])
        
        # 构建 Prompt
        prompt = TASK_DECOMPOSE_PROMPT.format(
            content=task.description,
            additional_info=additional_info or "",
            child_nodes_info=child_nodes_info
        )
        
        # 调用 LLM
        task_obj = Task(
            description=prompt,
            expected_output="XML format with <tasks> root containing <task> elements"
        )
        
        result = self.executor.run(self.task_agent, task_obj)
        
        # 解析结果
        output = result.get("result", result.get("output", ""))
        if isinstance(output, dict):
            output = str(output.get("content", output))
        
        # 解析 XML 格式的子任务
        subtasks = self._parse_subtasks_xml(output, parent_task_id=task.id, fallback_description=task.description)
        
        # 如果使用 AdaptivePlanner，可以进行优化
        if self.planner and subtasks:
            subtasks = await self._optimize_with_planner(task, subtasks)
        
        logger.info(f"[TaskDecomposer] Decomposed into {len(subtasks)} subtasks")
        return subtasks
    
    def _parse_subtasks_xml(
        self,
        xml_content: str,
        parent_task_id: str,
        fallback_description: Optional[str] = None,
    ) -> List[Task]:
        """解析 XML 格式的子任务"""
        subtasks = []
        
        try:
            # 尝试解析 XML
            root = ET.fromstring(f"<root>{xml_content}</root>")
            tasks_elem = root.find(".//tasks")
            
            if tasks_elem is not None:
                for i, task_elem in enumerate(tasks_elem.findall("task")):
                    task_text = task_elem.text or ""
                    if task_text.strip():
                        subtask = Task(
                            id=f"{parent_task_id}_subtask_{i+1}",
                            description=task_text.strip(),
                            expected_output="Task execution result",
                            dependencies=[],
                        )
                        subtasks.append(subtask)
            else:
                # 如果没有找到 <tasks>，尝试直接查找 <task>
                for i, task_elem in enumerate(root.findall(".//task")):
                    task_text = task_elem.text or ""
                    if task_text.strip():
                        subtask = Task(
                            id=f"{parent_task_id}_subtask_{i+1}",
                            description=task_text.strip(),
                            expected_output="Task execution result",
                            dependencies=[],
                        )
                        subtasks.append(subtask)
        except ET.ParseError:
            # XML 解析失败，尝试正则表达式提取
            logger.warning("[TaskDecomposer] XML parsing failed, using regex fallback")
            pattern = r"<task>(.*?)</task>"
            matches = re.findall(pattern, xml_content, re.DOTALL)
            
            for i, match in enumerate(matches):
                task_text = match.strip()
                if task_text:
                    subtask = Task(
                        id=f"{parent_task_id}_subtask_{i+1}",
                        description=task_text,
                        expected_output="Task execution result",
                        dependencies=[],
                    )
                    subtasks.append(subtask)
        
        # 如果没有解析到任何子任务，创建一个包含原始任务的子任务
        if not subtasks:
            logger.warning("[TaskDecomposer] No subtasks parsed, creating single subtask")
            subtasks.append(Task(
                id=f"{parent_task_id}_subtask_1",
                description=fallback_description or "Task execution",
                expected_output="Task execution result",
                dependencies=[],
            ))
        
        return subtasks
    
    async def _optimize_with_planner(
        self,
        parent_task: Task,
        subtasks: List[Task],
    ) -> List[Task]:
        """使用 AdaptivePlanner 优化子任务列表"""
        if not self.planner:
            return subtasks
        
        try:
            # 这里可以调用 AdaptivePlanner 进行优化
            # 例如：检查依赖关系、合并相似任务、调整优先级等
            # 目前先返回原始列表，后续可以扩展
            logger.debug("[TaskDecomposer] Using AdaptivePlanner for optimization")
            return subtasks
        except Exception as e:
            logger.warning(f"[TaskDecomposer] Planner optimization failed: {e}")
            return subtasks
    
    async def decompose_task_structured(
        self,
        task: Task,
        available_workers: List[Any],
        additional_info: Optional[str] = None,
    ) -> TaskDecompositionResult:
        """
        分解任务为结构化结果（使用 Pydantic 模型）
        
        Args:
            task: 要分解的任务
            available_workers: 可用的 Worker 列表
            additional_info: 额外信息（可选）
            
        Returns:
            TaskDecompositionResult: 结构化分解结果
        """
        # 先使用 XML 格式分解
        subtask_tasks = await self.decompose_task(
            task=task,
            available_workers=available_workers,
            additional_info=additional_info,
        )
        
        # 转换为结构化格式
        subtask_definitions = []
        for i, subtask_task in enumerate(subtask_tasks):
            # 提取依赖关系
            dependencies = subtask_task.dependencies or []
            
            subtask_def = SubtaskDefinition(
                description=subtask_task.description,
                expected_output=subtask_task.expected_output or "Task execution result",
                dependencies=dependencies,
                priority=1,  # 默认优先级
            )
            subtask_definitions.append(subtask_def)
        
        # 分析是否可以并行执行
        can_parallelize = all(
            len(subtask_def.dependencies) == 0
            for subtask_def in subtask_definitions
        )
        
        return TaskDecompositionResult(
            subtasks=subtask_definitions,
            reasoning=f"Decomposed task '{task.id}' into {len(subtask_definitions)} subtasks",
            can_parallelize=can_parallelize,
        )
    
    def collect_resource_config_interactive(self) -> ResourceConfig:
        """
        交互式收集资源配置（v3.3 新增）
        
        通过问答方式收集项目资源配置信息，用于甘特图演算。
        
        Returns:
            ResourceConfig: 资源配置对象
        """
        print("\n" + "="*60)
        print("📋 AgenticX TaskDecomposer v3.3 - 资源配置收集")
        print("="*60)
        
        # Q1: 后端人数
        backend_input = input("\n1️⃣  后端投入几人？ [直接回复数字，或\"不确定\"]: ").strip()
        backend_count = None if backend_input in ["不确定", "不知道", ""] else int(backend_input)
        
        # Q2: 后端可用时间
        if backend_count:
            backend_avail = input("   后端每人可用时间？ [默认 100%，直接回车跳过]: ").strip()
            backend_availability = float(backend_avail.replace("%", "")) if backend_avail else 100.0
        else:
            backend_availability = 100.0
        
        # Q3: 测试人数
        test_input = input("\n2️⃣  测试投入几人？ [直接回复数字，或\"不确定\"]: ").strip()
        test_count = None if test_input in ["不确定", "不知道", ""] else int(test_input)
        
        # Q4: 测试可用时间
        if test_count:
            test_avail = input("   测试每人可用时间？ [默认 100%，直接回车跳过]: ").strip()
            test_availability = float(test_avail.replace("%", "")) if test_avail else 100.0
        else:
            test_availability = 100.0
        
        # Q5: 期望完工日期
        target_date = input("\n3️⃣  有期望完工日期吗？ [可选，如\"2026-06-30\"，直接回车跳过]: ").strip()
        if not target_date:
            target_date = None
        
        # Q6: 固定里程碑
        milestones_input = input("4️⃣  有固定里程碑吗？ [可选，用逗号分隔，直接回车跳过]: ").strip()
        milestones = [m.strip() for m in milestones_input.split(",")] if milestones_input else None
        
        # Q7: 技能水平
        skill_input = input("5️⃣  团队技能水平？ [average/senior/mixed，默认 average]: ").strip().lower()
        skill_level = skill_input if skill_input in ["average", "senior", "mixed"] else "average"
        
        # Q8: 其他角色
        frontend_count = input("6️⃣  前端投入几人？ [默认 0]: ").strip()
        frontend_count = int(frontend_count) if frontend_count else 0
        
        product_count = input("7️⃣  产品几人？ [默认 0]: ").strip()
        product_count = int(product_count) if product_count else 0
        
        # 创建配置对象
        self.resource_config = ResourceConfig(
            backend_count=backend_count,
            backend_availability=backend_availability,
            test_count=test_count,
            test_availability=test_availability,
            frontend_count=frontend_count,
            product_count=product_count,
            target_date=target_date,
            milestones=milestones,
            skill_level=skill_level,
        )
        
        print("\n" + "="*60)
        print("✅ 资源配置已收集完成！")
        print("="*60)
        
        return self.resource_config
    
    def set_resource_config(self, config_dict: Dict[str, Any]) -> ResourceConfig:
        """
        通过字典设置资源配置（编程方式）
        
        Args:
            config_dict: 配置字典，键名同 ResourceConfig 字段
            
        Returns:
            ResourceConfig: 资源配置对象
        """
        self.resource_config = ResourceConfig(**config_dict)
        return self.resource_config
    
    def generate_scenarios(self) -> List[Dict[str, Any]]:
        """
        生成多场景对比（当人数不确定时）
        
        Returns:
            场景列表，每个场景包含配置和预估工期
        """
        scenarios = []
        
        # 基础场景：2 后端 +1 测试
        scenarios.append({
            "name": "精简版",
            "config": ResourceConfig(
                backend_count=2,
                backend_availability=100.0,
                test_count=1,
                test_availability=100.0,
                skill_level="average",
            ),
            "estimated_weeks": "12 周",
            "notes": "适合小团队，工期较长但人员压力小",
        })
        
        # 标准场景：3 后端 +1 测试
        scenarios.append({
            "name": "标准版",
            "config": ResourceConfig(
                backend_count=3,
                backend_availability=100.0,
                test_count=1,
                test_availability=100.0,
                skill_level="average",
            ),
            "estimated_weeks": "9 周",
            "notes": "推荐配置，工期和负载平衡",
        })
        
        # 加速场景：4 后端 +2 测试
        scenarios.append({
            "name": "加速版",
            "config": ResourceConfig(
                backend_count=4,
                backend_availability=100.0,
                test_count=2,
                test_availability=100.0,
                skill_level="senior",
            ),
            "estimated_weeks": "6 周",
            "notes": "适合紧急项目，需要更多人力资源",
        })
        
        return scenarios
    
    def calculate_gantt(
        self,
        tasks: List[Task],
        config: Optional[ResourceConfig] = None,
        start_date: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        基于资源配置计算甘特图（v3.4 新增）
        
        Args:
            tasks: 任务列表（Task 对象）
            config: 资源配置（可选，默认使用 self.resource_config）
            start_date: 开始日期（YYYY-MM-DD），默认今天
            
        Returns:
            甘特图数据字典，包含：
            - gantt_data: 甘特图数据列表
            - weekly_load: 周负载统计
            - total_weeks: 总工期（周）
            - start_date: 实际开始日期
            - end_date: 预计结束日期
        """
        config = config or self.resource_config
        
        if not config:
            logger.warning("[TaskDecomposer] No resource config, using default (3 backend + 1 test)")
            config = ResourceConfig(backend_count=3, test_count=1)
        
        # 设置开始日期
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
        else:
            start = datetime.now()
        
        # 计算每周可用工时
        backend_weekly_capacity = config.backend_count * 5 * (config.backend_availability / 100.0) if config.backend_count else 0
        test_weekly_capacity = config.test_count * 5 * (config.test_availability / 100.0) if config.test_count else 0
        
        logger.info(f"[TaskDecomposer] Calculating gantt with backend_capacity={backend_weekly_capacity} days/week, test_capacity={test_weekly_capacity} days/week")
        
        # 分析任务，估算工作量
        task_estimates = self._estimate_task_effort(tasks)
        
        # 分配任务到周次
        gantt_data, weekly_load = self._schedule_tasks(task_estimates, backend_weekly_capacity, test_weekly_capacity, start)
        
        # 计算总工期
        total_weeks = len(weekly_load)
        end_date = start + timedelta(weeks=total_weeks)
        
        return {
            "gantt_data": gantt_data,
            "weekly_load": weekly_load,
            "total_weeks": total_weeks,
            "start_date": start.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
            "config": config,
        }
    
    def _estimate_task_effort(self, tasks: List[Task]) -> List[Dict[str, Any]]:
        """
        估算任务工作量（人天）
        
        基于任务描述和复杂度进行估算。
        """
        estimates = []
        
        for i, task in enumerate(tasks):
            # 简单估算：根据描述长度和关键词
            desc = task.description.lower()
            
            # 默认 2 人天
            effort = 2.0
            
            # 根据关键词调整
            if any(kw in desc for kw in ["design", "架构", "设计", "database", "数据库"]):
                effort = 5.0
            elif any(kw in desc for kw in ["api", "interface", "接口", "endpoint"]):
                effort = 3.0
            elif any(kw in desc for kw in ["test", "测试", "qa", "validation"]):
                effort = 2.0
            elif any(kw in desc for kw in ["document", "文档", "readme"]):
                effort = 1.0
            
            # 根据依赖关系调整
            if task.dependencies:
                effort += len(task.dependencies) * 0.5  # 每个依赖增加 0.5 人天
            
            estimates.append({
                "task": task,
                "effort": effort,
                "role": "backend" if any(kw in desc for kw in ["test", "测试", "qa"]) else "backend",
                "dependencies": task.dependencies or [],
            })
        
        return estimates
    
    def _schedule_tasks(
        self,
        task_estimates: List[Dict[str, Any]],
        backend_capacity: float,
        test_capacity: float,
        start_date: datetime,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        调度任务到周次
        
        Returns:
            (gantt_data, weekly_load)
        """
        gantt_data = []
        weekly_load = []
        
        # 按依赖关系排序任务（简单的拓扑排序）
        sorted_tasks = self._topological_sort(task_estimates)
        
        # 任务完成时间追踪
        task_end_week = {}
        
        # 当前周工作量
        current_week = 0
        backend_load = 0.0
        test_load = 0.0
        
        for estimate in sorted_tasks:
            task = estimate["task"]
            effort = estimate["effort"]
            role = estimate["role"]
            
            # 检查依赖，确保所有依赖任务已完成
            min_start_week = 0
            for dep_id in estimate["dependencies"]:
                if dep_id in task_end_week:
                    min_start_week = max(min_start_week, task_end_week[dep_id])
            
            # 找到合适的周次
            week = min_start_week
            remaining_effort = effort
            
            while remaining_effort > 0:
                # 确保 weekly_load 有足够的数据
                while len(weekly_load) <= week:
                    weekly_load.append({
                        "week": week + 1,
                        "backend_load": 0.0,
                        "test_load": 0.0,
                        "tasks": [],
                    })
                
                # 计算本周可用容量
                capacity = backend_capacity if role == "backend" else test_capacity
                current_load = weekly_load[week]["backend_load"] if role == "backend" else weekly_load[week]["test_load"]
                available = capacity - current_load
                
                if available > 0:
                    # 可以安排任务
                    effort_this_week = min(remaining_effort, available)
                    
                    if role == "backend":
                        weekly_load[week]["backend_load"] += effort_this_week
                    else:
                        weekly_load[week]["test_load"] += effort_this_week
                    
                    weekly_load[week]["tasks"].append({
                        "task_id": task.id,
                        "task_desc": task.description[:50] + "..." if len(task.description) > 50 else task.description,
                        "effort": effort_this_week,
                        "role": role,
                    })
                    
                    remaining_effort -= effort_this_week
                    
                    if remaining_effort <= 0:
                        # 任务完成
                        task_end_week[task.id] = week + 1
                        
                        # 添加到甘特图数据
                        gantt_data.append({
                            "task_id": task.id,
                            "task_desc": task.description,
                            "role": role,
                            "start_week": min_start_week + 1,
                            "end_week": week + 1,
                            "effort": effort,
                            "dependencies": estimate["dependencies"],
                        })
                else:
                    # 本周容量不足，移到下一周
                    week += 1
                    if week > current_week:
                        current_week = week
        
        return gantt_data, weekly_load
    
    def _topological_sort(self, task_estimates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        简单的拓扑排序，确保依赖任务先执行
        """
        # 构建任务 ID 映射
        task_map = {est["task"].id: est for est in task_estimates}
        
        # 简单的排序：先排没有依赖的，再排有依赖的
        no_deps = [est for est in task_estimates if not est["dependencies"]]
        has_deps = [est for est in task_estimates if est["dependencies"]]
        
        return no_deps + has_deps
    
    def generate_excel(
        self,
        gantt_result: Dict[str, Any],
        output_path: str,
        project_name: str = "AgenticX Project",
    ) -> str:
        """
        生成 Excel 甘特图文件（v3.4 新增）
        
        Args:
            gantt_result: calculate_gantt() 的返回结果
            output_path: 输出文件路径
            project_name: 项目名称
            
        Returns:
            生成的文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("[TaskDecomposer] openpyxl not installed")
            raise ImportError("Please install openpyxl: pip install openpyxl")
        
        wb = Workbook()
        
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14, color="000000")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        
        # ========== Sheet 1: 甘特图总览 ==========
        ws_gantt = wb.active
        ws_gantt.title = "甘特图总览"
        
        # 标题
        ws_gantt.merge_cells("A1:H1")
        ws_gantt["A1"] = f"{project_name} - 甘特图总览"
        ws_gantt["A1"].font = title_font
        ws_gantt["A1"].alignment = center_align
        
        # 项目信息
        ws_gantt["A3"] = "项目名称:"
        ws_gantt["B3"] = project_name
        ws_gantt["A4"] = "开始日期:"
        ws_gantt["B4"] = gantt_result["start_date"]
        ws_gantt["A5"] = "结束日期:"
        ws_gantt["B5"] = gantt_result["end_date"]
        ws_gantt["A6"] = "总工期:"
        ws_gantt["B6"] = f"{gantt_result['total_weeks']} 周"
        ws_gantt["A7"] = "资源配置:"
        config = gantt_result["config"]
        ws_gantt["B7"] = f"后端{config.backend_count}人，测试{config.test_count}人"
        
        # 甘特图表格头
        headers = ["任务 ID", "任务描述", "角色", "开始周", "结束周", "工作量 (人天)", "依赖任务"]
        row = 9
        for col, header in enumerate(headers, 1):
            cell = ws_gantt.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # 甘特图数据
        for i, item in enumerate(gantt_result["gantt_data"], 1):
            row = 10 + i
            ws_gantt.cell(row=row, column=1, value=item["task_id"]).border = thin_border
            ws_gantt.cell(row=row, column=2, value=item["task_desc"][:80]).border = thin_border
            ws_gantt.cell(row=row, column=3, value=item["role"]).border = thin_border
            ws_gantt.cell(row=row, column=4, value=f"第{item['start_week']}周").border = thin_border
            ws_gantt.cell(row=row, column=5, value=f"第{item['end_week']}周").border = thin_border
            ws_gantt.cell(row=row, column=6, value=round(item["effort"], 1)).border = thin_border
            ws_gantt.cell(row=row, column=7, value=", ".join(item["dependencies"]) if item["dependencies"] else "-").border = thin_border
        
        # 调整列宽
        ws_gantt.column_dimensions["A"].width = 20
        ws_gantt.column_dimensions["B"].width = 50
        ws_gantt.column_dimensions["C"].width = 10
        ws_gantt.column_dimensions["D"].width = 10
        ws_gantt.column_dimensions["E"].width = 10
        ws_gantt.column_dimensions["F"].width = 12
        ws_gantt.column_dimensions["G"].width = 30
        
        # ========== Sheet 2: 周负载统计 ==========
        ws_load = wb.create_sheet(title="周负载统计")
        
        ws_load.merge_cells("A1:E1")
        ws_load["A1"] = f"{project_name} - 周负载统计"
        ws_load["A1"].font = title_font
        ws_load["A1"].alignment = center_align
        
        # 表格头
        headers = ["周次", "后端工作量 (人天)", "后端容量 (人天)", "测试工作量 (人天)", "测试容量 (人天)", "负载率"]
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws_load.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # 周负载数据
        backend_capacity = config.backend_count * 5 * (config.backend_availability / 100.0) if config.backend_count else 0
        test_capacity = config.test_count * 5 * (config.test_availability / 100.0) if config.test_count else 0
        
        for i, week_data in enumerate(gantt_result["weekly_load"], 1):
            row = 4 + i
            ws_load.cell(row=row, column=1, value=f"第{week_data['week']}周").border = thin_border
            
            backend_load = round(week_data["backend_load"], 1)
            test_load = round(week_data["test_load"], 1)
            
            ws_load.cell(row=row, column=2, value=backend_load).border = thin_border
            ws_load.cell(row=row, column=3, value=round(backend_capacity, 1)).border = thin_border
            ws_load.cell(row=row, column=4, value=test_load).border = thin_border
            ws_load.cell(row=row, column=5, value=round(test_capacity, 1)).border = thin_border
            
            # 计算负载率
            if backend_capacity > 0:
                load_rate = min(backend_load / backend_capacity, test_load / test_capacity if test_capacity > 0 else 0)
                ws_load.cell(row=row, column=6, value=f"{load_rate*100:.1f}%").border = thin_border
            else:
                ws_load.cell(row=row, column=6, value="N/A").border = thin_border
        
        ws_load.column_dimensions["A"].width = 10
        ws_load.column_dimensions["B"].width = 15
        ws_load.column_dimensions["C"].width = 15
        ws_load.column_dimensions["D"].width = 15
        ws_load.column_dimensions["E"].width = 15
        ws_load.column_dimensions["F"].width = 10
        
        # ========== Sheet 3: 资源配置 ==========
        ws_config = wb.create_sheet(title="资源配置")
        
        ws_config.merge_cells("A1:B1")
        ws_config["A1"] = f"{project_name} - 资源配置"
        ws_config["A1"].font = title_font
        ws_config["A1"].alignment = center_align
        
        config_data = [
            ("后端开发人数", config.backend_count),
            ("后端可用时间", f"{config.backend_availability}%"),
            ("测试人数", config.test_count),
            ("测试可用时间", f"{config.test_availability}%"),
            ("前端开发人数", config.frontend_count),
            ("产品人数", config.product_count),
            ("期望完工日期", config.target_date or "未设置"),
            ("团队技能水平", config.skill_level),
            ("固定里程碑", ", ".join(config.milestones) if config.milestones else "无"),
        ]
        
        for i, (label, value) in enumerate(config_data, 1):
            row = 3 + i
            ws_config.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_config.cell(row=row, column=1).border = thin_border
            ws_config.cell(row=row, column=2, value=value).border = thin_border
        
        ws_config.column_dimensions["A"].width = 20
        ws_config.column_dimensions["B"].width = 30
        
        # 保存文件
        wb.save(output_path)
        logger.info(f"[TaskDecomposer] Excel generated: {output_path}")
        
        return output_path

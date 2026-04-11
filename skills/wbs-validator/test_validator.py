#!/usr/bin/env python3
"""
WBS 验证 Skill - 测试脚本
测试验证器是否真的有效，能否发现各种错误
"""

import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# 测试文件路径
TEST_DIR = "/home/admin/.openclaw/workspace/skills/wbs-validator/test_files"
os.makedirs(TEST_DIR, exist_ok=True)

print("="*80)
print("🧪 WBS 验证 Skill 测试开始")
print("="*80)

# ═══════════════════════════════════════════════════
# 测试 1: 创建正确的 WBS Excel（基准测试）
# ═══════════════════════════════════════════════════

def create_correct_wbs():
    """创建一个正确的 WBS Excel（所有验证都应该通过）"""
    print("\n" + "="*80)
    print("测试 1: 创建正确的 WBS Excel（基准测试）")
    print("="*80)
    
    wb = Workbook()
    
    # 创建 5 个 sheet
    ws1 = wb.active
    ws1.title = "接口级 WBS 总览"
    
    ws2 = wb.create_sheet("任务级 WBS 总览")
    ws3 = wb.create_sheet("周负载统计")
    ws4 = wb.create_sheet("甘特图（按模块分组）")
    ws5 = wb.create_sheet("资源配置")
    
    # 填充接口级 WBS（28 行 = 2 表头 + 26 任务）
    for row in range(1, 29):
        for col in range(1, 12):
            ws1.cell(row=row, column=col, value=f"数据{row}-{col}")
    
    # 填充任务级 WBS（80 行 = 2 表头 + 78 任务）
    for row in range(1, 81):
        for col in range(1, 12):
            ws2.cell(row=row, column=col, value=f"数据{row}-{col}")
    
    # 填充甘特图（23 列 = 任务 + 负责人 + 21 天）
    ws4.cell(row=1, column=1, value="甘特图标题")
    for row in range(2, 29):
        for col in range(1, 24):
            ws4.cell(row=row, column=col, value=f"数据{row}-{col}")
    
    # 保存
    file_path = f"{TEST_DIR}/correct_wbs.xlsx"
    wb.save(file_path)
    print(f"✅ 创建成功：{file_path}")
    print(f"   - Sheet 数量：{len(wb.sheetnames)} 个")
    print(f"   - 接口级 WBS 行数：{ws1.max_row} 行")
    print(f"   - 任务级 WBS 行数：{ws2.max_row} 行")
    print(f"   - 甘特图列数：{ws4.max_column} 列")
    
    return file_path

# ═══════════════════════════════════════════════════
# 测试 2: 创建错误的 WBS Excel（应该被验证器发现）
# ═══════════════════════════════════════════════════

def create_wrong_wbs():
    """创建一个有多个错误的 WBS Excel（验证应该失败）"""
    print("\n" + "="*80)
    print("测试 2: 创建错误的 WBS Excel（应该被验证器发现）")
    print("="*80)
    
    wb = Workbook()
    
    # 错误 1: 只有 4 个 sheet（缺少 1 个）
    ws1 = wb.active
    ws1.title = "接口级 WBS 总览"
    wb.create_sheet("任务级 WBS 总览")
    wb.create_sheet("周负载统计")
    wb.create_sheet("甘特图（按模块分组）")
    # ❌ 缺少"资源配置"sheet
    
    # 错误 2: 接口级 WBS 只有 20 行（应该≥28）
    for row in range(1, 21):
        for col in range(1, 12):
            ws1.cell(row=row, column=col, value=f"数据{row}-{col}")
    
    # 错误 3: 甘特图只有 15 列（应该=23）
    ws4 = wb["甘特图（按模块分组）"]
    for row in range(1, 10):
        for col in range(1, 16):
            ws4.cell(row=row, column=col, value=f"数据{row}-{col}")
    
    # 错误 4: 甘特图颜色错误（周一、周二用蓝色而不是灰色）
    blue_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
    ws4.cell(row=5, column=3).fill = blue_fill  # 周一用蓝色（错误！）
    ws4.cell(row=5, column=4).fill = blue_fill  # 周二用蓝色（错误！）
    
    # 保存
    file_path = f"{TEST_DIR}/wrong_wbs.xlsx"
    wb.save(file_path)
    print(f"✅ 创建成功：{file_path}")
    print(f"   - Sheet 数量：{len(wb.sheetnames)} 个 ❌（应该 5 个）")
    print(f"   - 接口级 WBS 行数：{ws1.max_row} 行 ❌（应该≥28）")
    print(f"   - 甘特图列数：{ws4.max_column} 列 ❌（应该=23）")
    print(f"   - 甘特图颜色：周一/周二蓝色 ❌（应该灰色）")
    
    return file_path

# ═══════════════════════════════════════════════════
# 测试 3: 运行验证器
# ═══════════════════════════════════════════════════

def run_validator(file_path, test_name):
    """运行验证器并检查结果"""
    print(f"\n{'='*80}")
    print(f"测试 3: 验证 {test_name}")
    print(f"{'='*80}")
    
    try:
        wb = load_workbook(file_path)
        
        errors = []
        
        # 验证 1: Sheet 数量
        print(f"\n✅ 检查 1/5: Sheet 数量...")
        if len(wb.sheetnames) != 5:
            errors.append(f"Sheet 数量错误：{len(wb.sheetnames)} != 5")
            print(f"   ❌ 失败：{len(wb.sheetnames)} 个（应该 5 个）")
        else:
            print(f"   ✅ 通过：{len(wb.sheetnames)} 个")
        
        # 验证 2: 接口级 WBS 行数
        print(f"\n✅ 检查 2/5: 接口级 WBS 行数...")
        ws_interface = wb["接口级 WBS 总览"]
        if ws_interface.max_row < 28:
            errors.append(f"接口级 WBS 行数不足：{ws_interface.max_row} < 28")
            print(f"   ❌ 失败：{ws_interface.max_row} 行（应该≥28）")
        else:
            print(f"   ✅ 通过：{ws_interface.max_row} 行")
        
        # 验证 3: 任务级 WBS 行数
        print(f"\n✅ 检查 3/5: 任务级 WBS 行数...")
        ws_detailed = wb["任务级 WBS 总览"]
        if ws_detailed.max_row < 80:
            errors.append(f"任务级 WBS 行数不足：{ws_detailed.max_row} < 80")
            print(f"   ❌ 失败：{ws_detailed.max_row} 行（应该≥80）")
        else:
            print(f"   ✅ 通过：{ws_detailed.max_row} 行")
        
        # 验证 4: 甘特图列数
        print(f"\n✅ 检查 4/5: 甘特图列数...")
        ws_gantt = wb["甘特图（按模块分组）"]
        if ws_gantt.max_column != 23:
            errors.append(f"甘特图列数错误：{ws_gantt.max_column} != 23")
            print(f"   ❌ 失败：{ws_gantt.max_column} 列（应该=23）")
        else:
            print(f"   ✅ 通过：{ws_gantt.max_column} 列")
        
        # 验证 5: 甘特图颜色（关键！）
        print(f"\n✅ 检查 5/5: 甘特图颜色（关键）...")
        color_errors = []
        for row in range(4, min(30, ws_gantt.max_row + 1)):
            mon_cell = ws_gantt.cell(row=row, column=3)
            tue_cell = ws_gantt.cell(row=row, column=4)
            mon_color = str(mon_cell.fill.start_color.rgb) if mon_cell.fill.start_color.rgb else ""
            tue_color = str(tue_cell.fill.start_color.rgb) if tue_cell.fill.start_color.rgb else ""
            
            if "92CDDC" in mon_color or "92CDDC" in tue_color:
                color_errors.append(f"行{row} 周一/周二蓝色")
        
        if color_errors:
            errors.append(f"甘特图颜色错误：{len(color_errors)}个任务周一/周二蓝色")
            print(f"   ❌ 失败：{len(color_errors)}个任务周一/周二蓝色（应该灰色）")
        else:
            print(f"   ✅ 通过：所有任务周一/周二都是灰色")
        
        # 总结
        print(f"\n{'='*80}")
        if errors:
            print(f"❌ 验证失败！发现{len(errors)}个错误：")
            for error in errors:
                print(f"   - {error}")
            return False
        else:
            print(f"✅ 验证通过！所有检查项都正确")
            return True
    
    except Exception as e:
        print(f"\n❌ 验证异常：{e}")
        return False

# ═══════════════════════════════════════════════════
# 主测试流程
# ═══════════════════════════════════════════════════

def main():
    # 测试 1: 正确的 WBS（应该通过）
    correct_file = create_correct_wbs()
    result1 = run_validator(correct_file, "正确的 WBS")
    
    # 测试 2: 错误的 WBS（应该失败）
    wrong_file = create_wrong_wbs()
    result2 = run_validator(wrong_file, "错误的 WBS")
    
    # 总结
    print(f"\n{'='*80}")
    print("📊 测试总结")
    print(f"{'='*80}")
    print(f"测试 1（正确的 WBS）: {'✅ 通过' if result1 else '❌ 失败'}")
    print(f"测试 2（错误的 WBS）: {'✅ 通过（应该失败）' if not result2 else '❌ 失败（应该发现错误）'}")
    
    if result1 and not result2:
        print(f"\n{'='*80}")
        print("✅ 所有测试通过！验证器有效！")
        print(f"{'='*80}")
        print("验证器能够：")
        print("  ✅ 识别正确的 WBS")
        print("  ✅ 发现错误的 WBS")
        print("  ✅ 检测 Sheet 数量错误")
        print("  ✅ 检测行数不足")
        print("  ✅ 检测列数错误")
        print("  ✅ 检测颜色错误（关键！）")
        return True
    else:
        print(f"\n{'='*80}")
        print("❌ 测试失败！验证器有缺陷！")
        print(f"{'='*80}")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)

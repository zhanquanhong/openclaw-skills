"""Excel 输出模块"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import List, Dict, Optional
import re
import logging

logger = logging.getLogger(__name__)


def clean_text(text: str) -> str:
    """清理 Excel 不支持的字符"""
    if not text:
        return ''
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', str(text))
    return text


def _has_estimate(tasks: List[Dict]) -> bool:
    """检测任务列表是否包含估算字段"""
    return any('推荐人天数' in t for t in tasks)


def _append_summary_rows(ws, tasks: List[Dict], headers: List[str], cell_alignment, header_font, thin_border):
    """在数据行末尾追加模块小计和项目总计

    Args:
        ws: worksheet 对象
        tasks: 任务列表（含 推荐人天数 字段时计算小计）
        headers: 表头列表
        cell_alignment: 单元格样式
        header_font: 表头字体
        thin_border: 边框
    """
    if not _has_estimate(tasks):
        return

    # 按模块分组汇总
    module_totals: Dict[str, float] = {}
    module_counts: Dict[str, int] = {}
    for t in tasks:
        module = t.get('任务模块', '未分类')
        days = t.get('推荐人天数', 0)
        module_totals[module] = module_totals.get(module, 0) + days
        module_counts[module] = module_counts.get(module, 0) + 1

    # 空一行
    sep_row = len(tasks) + 2 + 1  # header(1) + data(N) + empty(1)
    start_row = sep_row + 1

    # 模块小计
    sort_order = [
        '用户管理', '权限系统', '配置管理', '数据迁移',
        '前端', '后端', '接口', '数据库',
    ]
    seen_modules = set()
    for module in sort_order:
        if module in module_totals and module not in seen_modules:
            _write_summary_row(ws, start_row, module, module_counts[module],
                               module_totals[module], headers, cell_alignment, header_font, thin_border)
            seen_modules.add(module)
            start_row += 1
    # 未排序的模块
    for module, total in sorted(module_totals.items()):
        if module not in seen_modules:
            _write_summary_row(ws, start_row, module, module_counts[module],
                               total, headers, cell_alignment, header_font, thin_border)
            seen_modules.add(module)
            start_row += 1

    # 汇总行之间的分隔线
    summary_font = Font(bold=True, size=11)

    # 项目总计
    grand_total = sum(module_totals.values())
    grand_count = sum(module_counts.values())
    start_row += 0  # no extra blank row

    # 总计行
    ws.cell(row=start_row, column=1, value='').border = thin_border
    ws.cell(row=start_row, column=2, value='').border = thin_border
    ws.cell(row=start_row, column=3, value='项目总计').font = summary_font
    ws.cell(row=start_row, column=3).alignment = cell_alignment
    ws.cell(row=start_row, column=3).border = thin_border
    # 任务数放在 '任务来源' 列（原第4列）
    ws.cell(row=start_row, column=4, value=f'{grand_count} 个任务').font = summary_font
    ws.cell(row=start_row, column=4).alignment = cell_alignment
    ws.cell(row=start_row, column=4).border = thin_border
    # 标记人天数的列
    estimate_col = headers.index('推荐人天数') + 1 if '推荐人天数' in headers else 10
    ws.cell(row=start_row, column=estimate_col, value=grand_total).font = summary_font
    ws.cell(row=start_row, column=estimate_col).alignment = cell_alignment
    ws.cell(row=start_row, column=estimate_col).border = thin_border
    ws.cell(row=start_row, column=estimate_col).number_format = '0.0'

    # 其余填充空
    for col in range(1, len(headers) + 1):
        if col not in (3, 4, estimate_col):
            ws.cell(row=start_row, column=col).border = thin_border


def _write_summary_row(ws, row, module, count, total, headers, cell_alignment, header_font, thin_border):
    """写一行模块小计"""
    summary_font = Font(bold=True, size=10)
    ws.cell(row=row, column=1, value='').border = thin_border
    ws.cell(row=row, column=2, value='').border = thin_border
    ws.cell(row=row, column=3, value=f'  {module} 小计').font = summary_font
    ws.cell(row=row, column=3).alignment = cell_alignment
    ws.cell(row=row, column=3).border = thin_border
    ws.cell(row=row, column=4, value=f'{count} 个任务').font = summary_font
    ws.cell(row=row, column=4).alignment = cell_alignment
    ws.cell(row=row, column=4).border = thin_border

    estimate_col = headers.index('推荐人天数') + 1 if '推荐人天数' in headers else 10
    ws.cell(row=row, column=estimate_col, value=total).font = summary_font
    ws.cell(row=row, column=estimate_col).alignment = cell_alignment
    ws.cell(row=row, column=estimate_col).border = thin_border
    ws.cell(row=row, column=estimate_col).number_format = '0.0'

    for col in range(1, len(headers) + 1):
        if col not in (3, 4, estimate_col):
            ws.cell(row=row, column=col).border = thin_border


def export_to_excel(
    tasks: List[Dict],
    output_path: str,
    validation_results: Optional[Dict[str, List]] = None,
    timing_info: Optional[Dict[str, float]] = None,
) -> None:
    """
    导出任务到 Excel，包含验证结果 sheet

    Args:
        tasks: 任务列表
        output_path: 输出文件路径
        validation_results: 验证结果（可选），用于"验证结果"sheet
        timing_info: 时序信息（可选），用于统计
    """
    wb = Workbook()

    # ==================== Sheet 1: 任务分解总表 ====================
    ws = wb.active
    ws.title = "任务分解总表"

    # 样式定义
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(vertical='center', wrap_text=True)

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头（v4.0 + v5.0 估算）
    has_estimate = _has_estimate(tasks)
    headers = ['任务模块', '任务 ID', '任务内容', '任务来源',
               '任务类型', '依赖', '可验收标准', '验证状态', '处理时间(ms)']
    if has_estimate:
        headers.extend(['推荐人天数', '估算范围'])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 验证状态颜色
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # 数据行
    for row_idx, task in enumerate(tasks, 2):
        ws.cell(row=row_idx, column=1, value=clean_text(task.get('任务模块', '未分类'))).alignment = cell_alignment
        ws.cell(row=row_idx, column=2, value=clean_text(task.get('任务 ID', ''))).alignment = cell_alignment
        ws.cell(row=row_idx, column=3, value=clean_text(task.get('任务内容', ''))).alignment = cell_alignment
        ws.cell(row=row_idx, column=4, value=clean_text(task.get('任务来源', ''))).alignment = cell_alignment

        ws.cell(row=row_idx, column=5, value=clean_text(task.get('任务类型', '普通任务'))).alignment = cell_alignment
        ws.cell(row=row_idx, column=6, value=clean_text(task.get('依赖', '无'))).alignment = cell_alignment
        ws.cell(row=row_idx, column=7, value=clean_text(task.get('可验收标准', ''))).alignment = cell_alignment

        # 验证状态列
        validation_status = task.get('验证状态', 'PASS')
        validation_cell = ws.cell(row=row_idx, column=8, value=validation_status)
        validation_cell.alignment = cell_alignment
        if validation_status == 'PASS':
            validation_cell.fill = green_fill
        elif validation_status == 'WARN':
            validation_cell.fill = yellow_fill
        elif validation_status == 'FAIL':
            validation_cell.fill = red_fill

        # 处理时间列
        proc_time = task.get('处理时间(ms)', '')
        ws.cell(row=row_idx, column=9, value=str(proc_time) if proc_time else '').alignment = cell_alignment

        # 估算列（如果有）
        if has_estimate:
            est_days = task.get('推荐人天数', '')
            if isinstance(est_days, (int, float)):
                est_cell = ws.cell(row=row_idx, column=10, value=est_days)
                est_cell.number_format = '0.0'
            else:
                ws.cell(row=row_idx, column=10, value='')
            est_cell.alignment = cell_alignment

            est_range = task.get('估算范围', '')
            ws.cell(row=row_idx, column=11, value=clean_text(est_range)).alignment = cell_alignment

        # 应用边框
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border

    # 调整列宽
    col_widths = {
        'A': 15, 'B': 10, 'C': 50, 'D': 40, 'E': 40,
        'F': 15, 'G': 10, 'H': 35, 'I': 12,
    }
    if has_estimate:
        col_widths['J'] = 14
        col_widths['K'] = 18
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 底部模块小计和总计行
    _append_summary_rows(ws, tasks, headers, cell_alignment, header_font, thin_border)

    # 自动筛选（有估算列时排除汇总行）
    data_end = 1 + len(tasks)
    ws.auto_filter.ref = f'A1:{chr(64 + len(headers))}{data_end}'

    # ==================== Sheet 2: 验证结果 ====================
    if validation_results:
        ws2 = wb.create_sheet("验证结果")

        # 表头
        v_headers = ['验证项', '状态', '任务内容', '详细信息']
        for col, header in enumerate(v_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        row_idx = 2
        for vtype, vlist in validation_results.items():
            if not vlist:
                continue
            for vresult in vlist:
                if isinstance(vresult, dict):
                    issues = vresult.get('issues', [])
                    status = 'PASS' if vresult.get('is_valid', True) else 'FAIL'
                    content = vresult.get('content', '')
                else:
                    issues = vresult.issues if hasattr(vresult, 'issues') else []
                    status = 'PASS' if vresult.is_valid else 'FAIL'
                    content = vresult.get('content', '') if isinstance(vresult, dict) else ''

                issue_text = '; '.join(issues) if issues else '无'
                ws2.cell(row=row_idx, column=1, value=vtype).alignment = cell_alignment
                ws2.cell(row=row_idx, column=2, value=status).alignment = cell_alignment
                ws2.cell(row=row_idx, column=3, value=clean_text(content)).alignment = cell_alignment
                ws2.cell(row=row_idx, column=4, value=clean_text(issue_text)).alignment = cell_alignment
                for col in range(1, 5):
                    ws2.cell(row=row_idx, column=col).border = thin_border

                status_cell = ws2.cell(row=row_idx, column=2)
                if status == 'FAIL':
                    status_cell.fill = red_fill
                elif status == 'WARN':
                    status_cell.fill = yellow_fill
                else:
                    status_cell.fill = green_fill

                row_idx += 1

        # 调整列宽
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 50
        ws2.column_dimensions['D'].width = 60

        # 时序信息（如果有）
        if timing_info:
            row_idx += 1
            ws2.cell(row=row_idx, column=1, value="性能统计").font = Font(bold=True, size=11)
            row_idx += 1
            for key, value in timing_info.items():
                ws2.cell(row=row_idx, column=1, value=key)
                ws2.cell(row=row_idx, column=2, value=f"{value:.1f}ms" if isinstance(value, float) else str(value))
                for col in range(1, 3):
                    ws2.cell(row=row_idx, column=col).border = thin_border
                row_idx += 1

    wb.save(output_path)
    logger.info(f'Excel 导出完成：{output_path}')
